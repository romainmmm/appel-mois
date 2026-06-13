import os
import tempfile
from datetime import date

import openpyxl
import pytest

from cleaning_schedule import CleaningTask
from staff import Worker
from distribution import assign_day
from excel_export import build_month_workbook, build_day_sheet


def _task(room, kind="depart"):
    return CleaningTask(room=room, floor=(room // 100) * 100, kind=kind, booking_id="b")


def _everyday(n):
    return {wd: n for wd in range(7)}


def _workers():
    return [
        Worker("Anna", 1, home_floor=100, floor_strict=True, weekly_max=_everyday(5)),
        Worker("Oumar", 3, weekly_max=_everyday(10)),
    ]


def _schedule():
    return {
        date(2026, 7, 6): [_task(101), _task(102), _task(301)],
        date(2026, 7, 7): [_task(103, "service")],
    }


class TestMonthWorkbook:
    def test_creates_file_with_dashboard(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            build_month_workbook(_schedule(), _workers(), path)
            wb = openpyxl.load_workbook(path)
            assert "Vue mensuelle" in wb.sheetnames
        finally:
            os.unlink(path)

    def test_dashboard_has_worker_columns(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            build_month_workbook(_schedule(), _workers(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Vue mensuelle"]
            header = [c.value for c in ws[1]]
            assert "Anna" in header
            assert "Oumar" in header
        finally:
            os.unlink(path)

    def test_has_one_sheet_per_day(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            build_month_workbook(_schedule(), _workers(), path)
            wb = openpyxl.load_workbook(path)
            # dashboard + 2 day sheets
            assert len(wb.sheetnames) == 3
        finally:
            os.unlink(path)


class TestDaySheet:
    def test_day_sheet_lists_rooms(self):
        day = date(2026, 7, 6)
        da = assign_day(_schedule()[day], _workers(), day)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            build_day_sheet(da, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # flatten all cell text
            text = " ".join(
                str(c.value) for row in ws.iter_rows() for c in row if c.value
            )
            assert "Anna" in text
            assert "101" in text
        finally:
            os.unlink(path)
