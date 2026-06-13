import os
import tempfile
from datetime import date

import openpyxl

from timesheet import set_entry, fortnight
from excel_export import build_timesheet_workbook


def _build(path):
    ts = {}
    set_entry(ts, "Anna", "2026-06-15", "08:00", "16:00", 30)   # 7.5
    set_entry(ts, "Anna", "2026-06-16", "09:00", "17:00", 0)    # 8.0
    set_entry(ts, "Sophie", "2026-06-15", "07:00", "15:00", 0)  # 8.0
    dates = fortnight(date(2026, 6, 15))
    employees = [("Anna", "Équipe ménage"), ("Sophie", "Accueil")]
    build_timesheet_workbook(dates, employees, ts, path)


def test_creates_file_with_sheets():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        _build(path)
        wb = openpyxl.load_workbook(path)
        assert "Récapitulatif paie" in wb.sheetnames
        assert "Détail par jour" in wb.sheetnames
    finally:
        os.unlink(path)


def test_totals_correct():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        _build(path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Récapitulatif paie"]
        # flatten
        text = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:
                text[row[0]] = row[2]
        assert text["Anna"] == 15.5
        assert text["Sophie"] == 8.0
    finally:
        os.unlink(path)
