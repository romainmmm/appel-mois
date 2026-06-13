import os
import tempfile
import pytest
import openpyxl
from pdf_parser import RoomEntry
from excel_generator import generate_excel
from room_layout import ROOM_LAYOUT


def _make_data():
    return {
        "date": "Mercredi 16 décembre 2026",
        "arrivees": [
            RoomEntry(102, "nom-test prenom-test", ""),
            RoomEntry(401, "nom-test prenom-test", ""),
            RoomEntry(101, "nom-test prenom-test", ""),
            RoomEntry(210, "nom-test prenom-test", ""),
        ],
        "departs": [
            RoomEntry(210, "nom-test prenom-test", ""),
        ],
        "service": [
            RoomEntry(201, "nom-test prenom-test", "Nuit 1 sur 3"),
            RoomEntry(211, "nom-test prenom-test", "Nuit 1 sur 3"),
        ],
    }


class TestGenerateExcel:
    def test_creates_file(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_date_in_header(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert "Mercredi 16 décembre 2026" in str(ws["A1"].value)
        finally:
            os.unlink(path)

    def test_room_numbers_present(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Check a few room numbers are in their expected cells
            row, num_col, _ = ROOM_LAYOUT[218]
            assert ws[f"{num_col}{row}"].value == 218
            row, num_col, _ = ROOM_LAYOUT[101]
            assert ws[f"{num_col}{row}"].value == 101
            row, num_col, _ = ROOM_LAYOUT[414]
            assert ws[f"{num_col}{row}"].value == 414
        finally:
            os.unlink(path)

    def test_arrivee_marked(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row, _, info_col = ROOM_LAYOUT[102]
            cell_value = ws[f"{info_col}{row}"].value
            assert cell_value is not None
            assert "ARR" in cell_value.upper() or "ARRIV" in cell_value.upper()
        finally:
            os.unlink(path)

    def test_depart_marked(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row, _, info_col = ROOM_LAYOUT[210]
            cell_value = ws[f"{info_col}{row}"].value
            assert cell_value is not None
            assert "DÉP" in cell_value.upper() or "DEP" in cell_value.upper()
        finally:
            os.unlink(path)

    def test_service_marked(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row, _, info_col = ROOM_LAYOUT[201]
            cell_value = ws[f"{info_col}{row}"].value
            assert cell_value is not None
            assert "SERVICE" in cell_value.upper() or "NUIT" in cell_value.upper()
        finally:
            os.unlink(path)

    def test_empty_room_is_blank(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Room 318 has no task
            row, _, info_col = ROOM_LAYOUT[318]
            cell_value = ws[f"{info_col}{row}"].value
            assert cell_value is None or str(cell_value).strip() == ""
        finally:
            os.unlink(path)

    def test_turnover_room_shows_both(self):
        """Room 210 is both departure and arrival — should show both."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            generate_excel(_make_data(), path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row, _, info_col = ROOM_LAYOUT[210]
            cell_value = ws[f"{info_col}{row}"].value.upper()
            assert "DÉP" in cell_value or "DEP" in cell_value
            assert "ARR" in cell_value
        finally:
            os.unlink(path)
