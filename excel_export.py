"""Generate the Excel deliverables.

Two products:
  - build_month_workbook : a 'Vue mensuelle' dashboard + one day-grid sheet per day.
  - build_day_sheet       : a single printable day-grid sheet.

The day grid reproduces the motel's paper day sheet (fixed room positions by
floor); each occupied room shows the cleaning type and the assigned worker,
e.g. "DÉPART | Anna" or "SERVICE | Oumar (Nuit 3)".
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from cleaning_schedule import CleaningTask
from staff import Worker, WEEKDAYS_FR
from distribution import assign_day, DayAssignment
from room_layout import ROOM_LAYOUT
import palette


# Fills (sober palette)
_FILL_DEPART = PatternFill("solid", fgColor=palette.DEPART)
_FILL_SERVICE = PatternFill("solid", fgColor=palette.SERVICE)
_FILL_MANUAL = PatternFill("solid", fgColor=palette.MANUAL)
_FILL_MANAGER = PatternFill("solid", fgColor=palette.MANAGER)
_FILL_HEADER = PatternFill("solid", fgColor=palette.HEADER)
_FILL_WEEKEND = PatternFill("solid", fgColor=palette.WEEKEND)

_THIN = Side(style="thin", color=palette.GRID)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ROW_HEIGHT = 35
_FS_ROOM = 14
_FS_INFO = 12
_FS_HEADER = 20

MANAGER_LABEL = "Gérants (à replanifier)"


def _safe_sheet_name(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def _setup_print(ws, landscape=True):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)


# ── Monthly dashboard ───────────────────────────────────────────────
def build_month_workbook(
    schedule: dict[date, list[CleaningTask]],
    workers: list[Worker],
    output_path: str,
) -> None:
    ordered = sorted(workers, key=lambda w: w.order)
    days = sorted(schedule.keys())

    wb = Workbook()
    dash = wb.active
    dash.title = "Vue mensuelle"
    dash.page_setup.orientation = "landscape"
    dash.page_setup.fitToWidth = 1
    dash.sheet_properties.pageSetUpPr.fitToPage = True

    headers = ["Date", "Jour", "Total"] + [w.name for w in ordered] + [MANAGER_LABEL]
    for col, label in enumerate(headers, start=1):
        c = dash.cell(row=1, column=col, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER

    for r, d in enumerate(days, start=2):
        da = assign_day(schedule[d], ordered, d)
        is_weekend = d.weekday() >= 5
        row_vals = [d.strftime("%d/%m/%Y"), WEEKDAYS_FR[d.weekday()], len(schedule[d])]
        for w in ordered:
            row_vals.append(len(da.assignments.get(w.name, [])) or "")
        row_vals.append(len(da.unassigned) or "")
        for col, val in enumerate(row_vals, start=1):
            c = dash.cell(row=r, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _BORDER
            if is_weekend:
                c.fill = _FILL_WEEKEND
            if col == len(row_vals) and val:
                c.fill = _FILL_MANAGER
                c.font = Font(bold=True)

    dash.column_dimensions["A"].width = 12
    dash.column_dimensions["B"].width = 11
    dash.column_dimensions["C"].width = 7
    for i in range(len(ordered) + 1):
        dash.column_dimensions[get_column_letter(4 + i)].width = 12
    dash.freeze_panes = "A2"

    for d in days:
        da = assign_day(schedule[d], ordered, d)
        ws = wb.create_sheet(title=_safe_sheet_name(d))
        _write_day_grid(ws, da)

    wb.save(output_path)


# ── Single day sheet ────────────────────────────────────────────────
def build_day_sheet(day_assignment: DayAssignment, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(day_assignment.day)
    _write_day_grid(ws, day_assignment)
    wb.save(output_path)


def _room_cell_text(tasks: list[CleaningTask], worker: str) -> tuple[str, PatternFill]:
    """Build the info-cell text + fill for a room given its task(s) and worker."""
    kinds = {t.kind for t in tasks}

    # Manual-only task (Ménage/Serviette added by hand)
    if kinds == {"manuel"}:
        label = " ; ".join(t.night_label for t in tasks if t.night_label) or "Manuel"
        fill = _FILL_MANUAL
        text = f"{label} | {worker}"
        if worker == MANAGER_LABEL:
            fill = _FILL_MANAGER
        return text, fill

    # Automatic task(s), possibly combined with a manual one
    if "depart" in kinds and "service" in kinds:
        label, fill = "DÉP+SERV", _FILL_DEPART
    elif "depart" in kinds:
        label, fill = "DÉPART", _FILL_DEPART
    elif "service" in kinds:
        label, fill = "SERVICE", _FILL_SERVICE
    else:
        label, fill = "Manuel", _FILL_MANUAL
    night = next((t.night_label for t in tasks if t.kind == "service" and t.night_label), "")
    text = f"{label} | {worker}"
    if night:
        text += f" ({night})"
    if "manuel" in kinds:
        extra = " ; ".join(t.night_label for t in tasks if t.kind == "manuel" and t.night_label)
        if extra:
            text += f" + {extra}"
    if worker == MANAGER_LABEL:
        fill = _FILL_MANAGER
    return text, fill


def _write_day_grid(ws, da: DayAssignment) -> None:
    _setup_print(ws, landscape=True)
    d = da.day

    # Map each room -> (tasks, worker_name)
    room_worker: dict[int, str] = {}
    room_tasks: dict[int, list[CleaningTask]] = {}
    for name, tasks in da.assignments.items():
        for t in tasks:
            room_worker[t.room] = name
            room_tasks.setdefault(t.room, []).append(t)
    for t in da.unassigned:
        room_worker[t.room] = MANAGER_LABEL
        room_tasks.setdefault(t.room, []).append(t)

    # Title
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"Feuille du jour — {WEEKDAYS_FR[d.weekday()]} {d.strftime('%d/%m/%Y')}"
    title.font = Font(bold=True, size=_FS_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[11].height = 8  # separator between zones

    for room, (row, num_col, info_col) in ROOM_LAYOUT.items():
        ws.row_dimensions[row].height = _ROW_HEIGHT

        num_cell = ws[f"{num_col}{row}"]
        num_cell.value = room
        num_cell.font = Font(bold=True, size=_FS_ROOM)
        num_cell.border = _BORDER
        num_cell.alignment = Alignment(horizontal="center", vertical="center")

        info_cell = ws[f"{info_col}{row}"]
        info_cell.border = _BORDER
        info_cell.alignment = Alignment(wrap_text=False, vertical="center", horizontal="left")
        info_cell.font = Font(size=_FS_INFO)

        if room in room_tasks:
            text, fill = _room_cell_text(room_tasks[room], room_worker[room])
            info_cell.value = text
            info_cell.fill = fill
            info_cell.font = Font(bold=True, size=_FS_INFO)

    # Column widths
    for col in ("A", "C", "E"):
        ws.column_dimensions[col].width = 7.5
    for col in ("B", "D", "F"):
        ws.column_dimensions[col].width = 50
    ws.column_dimensions["G"].width = 1

    # Legend
    legend_row = 23
    ws.row_dimensions[legend_row].height = 24
    ws[f"A{legend_row}"] = "Légende :"
    ws[f"A{legend_row}"].font = Font(bold=True, size=11)
    for col, label, fill in [
        ("B", "DÉPART", _FILL_DEPART),
        ("C", "SERVICE", _FILL_SERVICE),
        ("D", "Manuel", _FILL_MANUAL),
        ("E", "Gérants", _FILL_MANAGER),
    ]:
        cell = ws[f"{col}{legend_row}"]
        cell.value = label
        cell.fill = fill
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # Rooms not in the fixed layout (safety) -> note below the legend
    extras = [r for r in room_tasks if r not in ROOM_LAYOUT]
    if extras:
        ws[f"A{legend_row + 1}"] = "Chambres hors plan : " + ", ".join(str(r) for r in sorted(extras))
        ws[f"A{legend_row + 1}"].font = Font(italic=True, color="C00000")


# ── Timesheet / payroll workbook ────────────────────────────────────
def build_timesheet_workbook(dates, employees, timesheet_data, output_path):
    """Payroll workbook for a pay period.

    Args:
        dates: list of date objects covering the period (e.g. 14 days).
        employees: list of (name, role) tuples.
        timesheet_data: the timesheet dict {name: {iso: {arrivee, depart, pause}}}.
        output_path: .xlsx destination.
    """
    from timesheet import worked_hours, period_total

    wb = Workbook()

    # ── Sheet 1: payroll summary ────────────────────────────────────
    summary = wb.active
    summary.title = "Récapitulatif paie"
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_setup.fitToWidth = 1

    period = f"Quinzaine du {dates[0].strftime('%d/%m/%Y')} au {dates[-1].strftime('%d/%m/%Y')}"
    summary.merge_cells("A1:C1")
    summary["A1"] = period
    summary["A1"].font = Font(bold=True, size=14)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 26

    headers = ["Employé", "Rôle", "Total heures"]
    for col, label in enumerate(headers, start=1):
        c = summary.cell(row=2, column=col, value=label)
        c.font = Font(bold=True, color=palette.HEADER_TEXT, size=11)
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    grand_total = 0.0
    for r, (name, role) in enumerate(employees, start=3):
        total = period_total(timesheet_data, name, dates)
        grand_total += total
        for col, val in enumerate([name, role or "", total], start=1):
            c = summary.cell(row=r, column=col, value=val)
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="center" if col == 3 else "left", vertical="center")
        summary.cell(row=r, column=3).number_format = "0.00"

    tr = len(employees) + 3
    summary.cell(row=tr, column=2, value="TOTAL").font = Font(bold=True)
    tc = summary.cell(row=tr, column=3, value=grand_total)
    tc.font = Font(bold=True)
    tc.number_format = "0.00"
    tc.border = _BORDER

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 14

    # ── Sheet 2: daily detail ───────────────────────────────────────
    detail = wb.create_sheet("Détail par jour")
    detail.sheet_properties.pageSetUpPr.fitToPage = True
    detail.page_setup.fitToWidth = 1
    dheaders = ["Employé", "Date", "Jour", "Arrivée", "Départ", "Pause (min)", "Heures"]
    for col, label in enumerate(dheaders, start=1):
        c = detail.cell(row=1, column=col, value=label)
        c.font = Font(bold=True, color=palette.HEADER_TEXT, size=11)
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    from timesheet import get_entry
    row = 2
    for name, _role in employees:
        for d in dates:
            e = get_entry(timesheet_data, name, d.isoformat())
            h = worked_hours(e["arrivee"], e["depart"], e.get("pause", 0))
            if not e["arrivee"] and not e["depart"] and not h:
                continue  # skip empty days in the detail
            vals = [name, d.strftime("%d/%m/%Y"), WEEKDAYS_FR[d.weekday()],
                    e["arrivee"], e["depart"], e.get("pause", 0), h]
            for col, val in enumerate(vals, start=1):
                c = detail.cell(row=row, column=col, value=val)
                c.border = _BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
            detail.cell(row=row, column=7).number_format = "0.00"
            row += 1

    for col, w in zip("ABCDEFG", [20, 12, 10, 9, 9, 12, 9]):
        detail.column_dimensions[col].width = w

    wb.save(output_path)
