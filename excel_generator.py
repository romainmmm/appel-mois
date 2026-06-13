"""Generate the hotel housekeeping Excel sheet from parsed PDF data."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from room_layout import ROOM_LAYOUT, ALL_ROOMS
from pdf_parser import RoomEntry
import palette


# Color coding for task types (sober palette)
_FILL_ARRIVEE = PatternFill("solid", fgColor=palette.ARRIVEE)
_FILL_DEPART = PatternFill("solid", fgColor=palette.DEPART)
_FILL_SERVICE = PatternFill("solid", fgColor=palette.SERVICE)
_FILL_TURNOVER = PatternFill("solid", fgColor=palette.TURNOVER)

_THIN_BORDER = Border(
    left=Side(style="thin", color=palette.GRID),
    right=Side(style="thin", color=palette.GRID),
    top=Side(style="thin", color=palette.GRID),
    bottom=Side(style="thin", color=palette.GRID),
)

_ROW_HEIGHT = 35  # points — tall single line
_FONT_SIZE_ROOM = 14
_FONT_SIZE_INFO = 12
_FONT_SIZE_HEADER = 20


def generate_excel(data: dict, output_path: str) -> None:
    """
    Generate the housekeeping day sheet Excel file.

    Args:
        data: parsed PDF data with keys: date, arrivees, departs, service
        output_path: where to save the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (data["date"][:31] if data["date"] else "Feuille de jour")

    # ── Print / page setup ──────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.4, bottom=0.4,
        header=0.2, footer=0.2,
    )

    # Build lookup: room_number -> list of (type, entry)
    room_tasks = {}
    for entry in data["arrivees"]:
        room_tasks.setdefault(entry.room, []).append(("arrivee", entry))
    for entry in data["departs"]:
        room_tasks.setdefault(entry.room, []).append(("depart", entry))
    for entry in data["service"]:
        room_tasks.setdefault(entry.room, []).append(("service", entry))

    # ── Row 1: Date header (merged A1:G1) ───────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = data["date"]
    ws["A1"].font = Font(bold=True, size=_FONT_SIZE_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    # ── Separator row 11 (blank between zone 1 & 2) ────────────────
    ws.row_dimensions[11].height = 8

    # ── Write all room numbers and their info ───────────────────────
    for room_num, (row, num_col, info_col) in ROOM_LAYOUT.items():
        # Set row height
        ws.row_dimensions[row].height = _ROW_HEIGHT

        # Room number cell
        num_cell = ws[f"{num_col}{row}"]
        num_cell.value = room_num
        num_cell.font = Font(bold=True, size=_FONT_SIZE_ROOM)
        num_cell.border = _THIN_BORDER
        num_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Info cell — single line, no wrap
        info_cell = ws[f"{info_col}{row}"]
        info_cell.border = _THIN_BORDER
        info_cell.alignment = Alignment(
            wrap_text=False, vertical="center", horizontal="left"
        )
        info_cell.font = Font(size=_FONT_SIZE_INFO)

        if room_num in room_tasks:
            tasks = room_tasks[room_num]
            types = [t for t, _ in tasks]

            if "depart" in types and "arrivee" in types:
                # Turnover: departure + arrival same day
                text = "DÉP+ARR"
                names = set()
                for t, e in tasks:
                    names.add(e.name)
                if names:
                    text += f"  {' / '.join(names)}"
                info_cell.value = text
                info_cell.fill = _FILL_TURNOVER
                info_cell.font = Font(bold=True, size=_FONT_SIZE_INFO)
            elif "depart" in types:
                entry = [e for t, e in tasks if t == "depart"][0]
                info_cell.value = f"DÉPART  {entry.name}"
                info_cell.fill = _FILL_DEPART
                info_cell.font = Font(bold=True, size=_FONT_SIZE_INFO)
            elif "arrivee" in types:
                entry = [e for t, e in tasks if t == "arrivee"][0]
                info_cell.value = f"ARRIVÉE  {entry.name}"
                info_cell.fill = _FILL_ARRIVEE
                info_cell.font = Font(bold=True, size=_FONT_SIZE_INFO)
            elif "service" in types:
                entry = [e for t, e in tasks if t == "service"][0]
                extra = f" - {entry.extra}" if entry.extra else ""
                info_cell.value = f"SERVICE  {entry.name}{extra}"
                info_cell.fill = _FILL_SERVICE
                info_cell.font = Font(bold=True, size=_FONT_SIZE_INFO)

    # ── Column widths — large enough to fit on one line when printed ─
    ws.column_dimensions["A"].width = 7.5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 7.5
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 7.5
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 1   # thin spacer

    # ── Legend row ──────────────────────────────────────────────────
    legend_row = 23
    ws.row_dimensions[legend_row].height = 24
    ws[f"A{legend_row}"] = "Légende :"
    ws[f"A{legend_row}"].font = Font(bold=True, size=11)
    ws[f"A{legend_row}"].alignment = Alignment(vertical="center")

    for col, label, fill in [
        ("B", "ARRIVÉE", _FILL_ARRIVEE),
        ("C", "DÉPART", _FILL_DEPART),
        ("D", "SERVICE", _FILL_SERVICE),
        ("E", "DÉP + ARR", _FILL_TURNOVER),
    ]:
        cell = ws[f"{col}{legend_row}"]
        cell.value = label
        cell.fill = fill
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    wb.save(output_path)
